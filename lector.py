from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font

class Lector:
    def __init__(self):
        pass

    def obtener_novedades(self, ruta):
        wb = load_workbook(ruta)
        ws = wb.active
        result = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] != None:
                headers = [(ws.cell(row=3, column=col).value).strftime('%Y-%m-%d') for col in range(2, ws.max_column + 1)]

                for row in range(4, ws.max_row + 1):
                    name = ws.cell(row=row, column=1).value  # Obtener el nombre de la persona
                    if name:
                        result[name] = {}
                        for col, date in enumerate(headers, start=3):
                            status = ws.cell(row=row, column=col).value
                            result[name][date] = status
                    else:
                        break
                return result
            else:
                break

    def exportar_datos(self, marcas:list, marcas_horas:list, horasxnum):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Informe Marcas Reloj'

        # Escribir encabezados
        encabezados = ["N° Empleado", "Empleado", "Fecha", "Entrada", "Salida", "Novedad", "Observacion", "Hs. Trabajadas", "Función", "Sucursal", "Hs. Totales"]
        ws.append(encabezados)

        # Formateo de encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Copiar datos de los objetos Marca
        for marca in marcas:
            fila = [
                marca.n_empleado,
                marca.empleado,
                marca.fecha,
                marca.f_ingreso,
                marca.f_salida,
                marca.novedad,
                marca.observaciones,
                marca.hs_trabajadas,
                marca.funcion,
                marca.sucursal                
            ]
            ws.append(fila)
        persona = ''
        series_combinables = {}
        for row in ws.iter_rows(min_row=2):
            series_combinables[row[1].value] = []
        
        print(series_combinables)
        for row in ws.iter_rows():
            if row[0].row != 1 and persona == row[1].value:
                #ws.merge_cells(start_row=row[0].row-1, end_row=row[0].row, start_column=11, end_column=11)
                pass 
            else:
                persona = row[1].value

            if persona != 'Empleado':
                series_combinables[persona].append(row[0].row)
        horas_format = {}
        for value in horasxnum:
            horas_format[value['nombre']] = value['horas_trabajadas']


        for key, serie in series_combinables.items():
            ws.merge_cells(start_row=min(serie), end_row=max(serie), start_column=11, end_column=11)
            ws.cell(row=min(serie), column=11).value = horas_format[key]
        

        print(series_combinables)

        ws = wb.create_sheet(title='Reloj por semana')

         # Escribir encabezados
        encabezados = ["N° Empleado", "Empleado", "Sucursal", "Año", "Semana", "Horas Trabajadas"]
        ws.append(encabezados)

        # Formateo de encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Escribir datos
        for marca in marcas_horas:
            fila = [
                marca.n_empleado,
                marca.empleado,
                marca.sucursal,
                marca.año,
                marca.semana,
                marca.horas
            ]
            ws.append(fila)

        wb.save('datos_reloj.xlsx')
